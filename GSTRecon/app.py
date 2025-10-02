import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from fuzzywuzzy import fuzz
import plotly.express as px
import plotly.graph_objects as go
import logging
from utils.reconciliation import GSTReconciliation

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
from utils.helpers import format_currency, format_date, format_percentage, clean_gstin
from utils.settings import render_settings_page, apply_settings_to_reconciliation, get_current_settings
from utils.reports import render_unique_gst_report, get_report_summary, render_enhanced_insights, get_unmapped_gst_summary_for_excel, add_gstr2a_compliance_columns, get_gstr2a_compliance_summary, render_ai_insights_section_for_reports
from utils.merge_ledgers_ui import render_merge_ledgers_page
import io
import openpyxl
import time
from openpyxl.utils import get_column_letter
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode
from openpyxl.styles import PatternFill

def initialize_session_state():
    """Initialize session state variables if they don't exist."""
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'df' not in st.session_state:
        st.session_state.df = None
    if 'edited_df' not in st.session_state:
        st.session_state.edited_df = None
    if 'post_recon_comments' not in st.session_state:
        st.session_state.post_recon_comments = []
    if 'reconciliation' not in st.session_state:
        st.session_state.reconciliation = None
    if 'final_report' not in st.session_state:
        st.session_state.final_report = None
    if 'updated_report' not in st.session_state:
        st.session_state.updated_report = None
    if 'reconciliation_complete' not in st.session_state:
        st.session_state.reconciliation_complete = False
    if 'reconciliation_start_time' not in st.session_state:
        st.session_state.reconciliation_start_time = None
    if 'enhanced_run' not in st.session_state:
        st.session_state.enhanced_run = False
    if 'original_recon_summary' not in st.session_state:
        st.session_state.original_recon_summary = None
    if 'reconciliation_settings' not in st.session_state:
        st.session_state.reconciliation_settings = get_current_settings()

# Ensure session state is initialized before any logic
initialize_session_state()

# --- Custom CSS for modern look and table highlighting ---
st.markdown("""
    <style>
    body {
        font-family: 'Segoe UI', Arial, sans-serif;
        background-color: #f7f9fb;
    }
    .sidebar .sidebar-content {
        background: #1a237e;
        color: #fff;
    }
    .stApp {
        background-color: #f7f9fb;
    }
    .metric-card {
        background: #fff;
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        padding: 1.5rem;
        margin: 0.5rem 0;
        text-align: center;
    }
    .metric-value {
        font-size: 2rem;
        font-weight: bold;
        color: #1976d2;
    }
    .metric-label {
        font-size: 1rem;
        color: #666;
    }
    .status-badge {
        display: inline-block;
        padding: 0.25em 0.75em;
        border-radius: 12px;
        font-size: 0.9em;
        font-weight: 600;
        color: #fff;
    }
    .badge-matched { background: #43a047; }
    .badge-partial { background: #fbc02d; color: #222; }
    .badge-mismatch { background: #e53935; }
    .badge-eligible { background: #1976d2; }
    .badge-ignore { background: #757575; }
    .highlight-partial-row {
        background-color: #fffbe6 !important;
    }
    .highlight-diff-cell {
        background-color: #ffe0b2 !important;
        font-weight: bold;
    }
    .highlight-status-cell {
        background-color: #e3f2fd !important;
        font-weight: bold;
    }
    .highlight-suggestions-cell {
        background-color: #f1f8e9 !important;
    }
    .block-container {
        padding-left: 0rem !important;
        padding-right: 0rem !important;
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        max-width: 100vw !important;
    }
    .main .block-container {
        max-width: 100vw !important;
    }
    .ag-theme-streamlit .ag-root-wrapper {
        border-radius: 0 !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- Sidebar Navigation ---
st.sidebar.image("https://img.icons8.com/ios-filled/50/1976d2/tax.png", width=40)
st.sidebar.title("GST Reconciliation")
st.sidebar.markdown("---")
section = st.sidebar.radio(
    "Navigation",
    ["Dashboard", "GSTR1", "GSTR2A", "GSTR2B", "Merge Books Ledgers", "Reports", "Data Overview", "Settings", "Performance Monitor", "Help"],
    index=0,
    format_func=lambda x: f"📊 {x}" if x == "Dashboard" else f"⚙️ {x}" if x == "Settings" else f"🔄 {x}" if x == "Merge Books Ledgers" else x
)

# --- Header Bar ---
col_logo, col_title, col_period, col_actions = st.columns([1, 3, 2, 2])
with col_logo:
    st.image("https://img.icons8.com/ios-filled/50/1976d2/tax.png", width=40)
with col_title:
    st.markdown("<h2 style='margin-bottom:0;'>GST Reconciliation App</h2>", unsafe_allow_html=True)
with col_period:
    period = st.selectbox("Period", ["Apr 2024 - Mar 2025", "Apr 2023 - Mar 2024", "Custom"], key="period_select")
with col_actions:
    st.button("🔄 Refresh", key="refresh_btn")
    st.button("⬇️ Download", key="download_btn")

# --- File Upload and Reconciliation Trigger (always visible at top) ---
st.markdown("---")
st.header("Upload & Reconcile Data")

# Check if merged data is available
merged_data_available = st.session_state.get('merged_data_for_reconciliation') is not None

col1, col2 = st.columns([3, 1])
with col1:
    uploaded_file = st.file_uploader("Upload Excel File", type=['xlsx', 'xls'])
with col2:
    if merged_data_available:
        if st.button("📊 Use Merged Data", type="secondary", key="use_merged_data_btn"):
            st.session_state.uploaded_file = "merged_data"
            st.session_state.df = st.session_state.merged_data_for_reconciliation.copy()
            st.session_state.edited_df = st.session_state.df.copy()
            st.session_state.reconciliation = None
            st.session_state.final_report = None
            st.session_state.updated_report = None
            st.session_state.reconciliation_complete = False
            st.session_state.reconciliation_start_time = None
            st.session_state.enhanced_run = False
            st.session_state.original_recon_summary = None
            st.success("✅ Using merged data for reconciliation!")
            st.rerun()
        st.info("💡 Merged data available from 'Merge Books Ledgers'")

if uploaded_file is not None and uploaded_file != st.session_state.uploaded_file:
    st.session_state.uploaded_file = uploaded_file
    st.session_state.df = None
    st.session_state.edited_df = None
    st.session_state.reconciliation = None
    st.session_state.final_report = None
    st.session_state.updated_report = None
    st.session_state.reconciliation_complete = False
    st.session_state.reconciliation_start_time = None
    st.session_state.enhanced_run = False
    st.session_state.original_recon_summary = None

if st.session_state.uploaded_file is not None:
    if st.session_state.df is None:
        if st.session_state.uploaded_file == "merged_data":
            # Use merged data
            st.session_state.df = st.session_state.merged_data_for_reconciliation.copy()
        else:
            # Read from uploaded file
            st.session_state.df = pd.read_excel(st.session_state.uploaded_file)
        st.session_state.edited_df = st.session_state.df.copy()
    st.subheader("Raw Data Preview")
    st.dataframe(st.session_state.edited_df, use_container_width=True)
    if st.button("Start Reconciliation", type="primary"):
        # --- Live progress bar, timer, and stage display ---
        progress_placeholder = st.empty()
        timer_placeholder = st.empty()
        stage_placeholder = st.empty()
        import time
        start_time = time.time()
        progress = 0
        stages = [
            "Initializing data...",
            "Matching invoices in same financial year...",
            "Matching invoices across years...",
            "Prioritizing tax value matches...",
            "Finalizing results..."
        ]
        for i, stage in enumerate(stages):
            progress = (i+1)/len(stages)
            progress_placeholder.progress(progress)
            stage_placeholder.write(f"Stage: {stage}")
            elapsed = int(time.time() - start_time)
            timer_placeholder.write(f"Elapsed time: {elapsed} seconds")
            time.sleep(0.5)  # Simulate stage (remove or adjust as needed)
        # --- Actual reconciliation ---
        stage_placeholder.write("Stage: Running reconciliation logic...")
        progress_placeholder.progress(0.95)
        timer_placeholder.write(f"Elapsed time: {int(time.time() - start_time)} seconds")
        reconciliation = GSTReconciliation(df=st.session_state.edited_df)
        st.session_state.reconciliation = reconciliation
        results = reconciliation.get_results()
        final_report = results['final_report']
        
        # Apply reconciliation settings to the results
        settings = get_current_settings()
        if settings.get('auto_apply_settings', True):
            final_report = apply_settings_to_reconciliation(final_report, settings)
            st.info("✅ Reconciliation settings applied to results.")
        
        st.session_state.final_report = final_report
        st.session_state.reconciliation_complete = True
        progress_placeholder.progress(1.0)
        total_time = int(time.time() - start_time)
        timer_placeholder.write(f"Total time spent: {total_time} seconds")
        stage_placeholder.write("Stage: Complete!")
        if total_time > 60:
            st.warning(f"Reconciliation took {total_time} seconds. For large files, consider optimizing your data or contact support.")
        st.success("Reconciliation completed!")

# --- Settings Page ---
if section == "Settings":
    st.markdown("---")
    render_settings_page()
    
    # Show current settings impact on existing reconciliation
    if st.session_state.get('reconciliation_complete') and st.session_state.get('final_report') is not None:
        st.markdown("---")
        st.markdown("### Apply Settings to Existing Reconciliation")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Apply Settings to Results", type="secondary"):
                settings = get_current_settings()
                updated_report = apply_settings_to_reconciliation(st.session_state.final_report, settings)
                st.session_state.final_report = updated_report
                st.success("✅ Settings applied to existing reconciliation results!")
                st.rerun()
        
        with col2:
            if st.button("📊 View Settings Impact", type="secondary"):
                settings = get_current_settings()
                original_report = st.session_state.final_report.copy()
                updated_report = apply_settings_to_reconciliation(original_report, settings)
                
                # Show comparison
                st.markdown("#### Settings Impact Comparison")
                
                # Count changes in Tax Diff Status
                if 'Tax Diff Status' in original_report.columns and 'Tax Diff Status' in updated_report.columns:
                    tax_changes = (original_report['Tax Diff Status'] != updated_report['Tax Diff Status']).sum()
                    st.metric("Tax Diff Status Changes", tax_changes)
                
                # Count changes in Date Status
                if 'Date Status' in original_report.columns and 'Date Status' in updated_report.columns:
                    date_changes = (original_report['Date Status'] != updated_report['Date Status']).sum()
                    st.metric("Date Status Changes", date_changes)
                
                # Show sample of changed rows
                if tax_changes > 0 or date_changes > 0:
                    st.markdown("#### Sample of Changed Rows")
                    changed_mask = (
                        (original_report['Tax Diff Status'] != updated_report['Tax Diff Status']) |
                        (original_report['Date Status'] != updated_report['Date Status'])
                    )
                    changed_rows = updated_report[changed_mask].head(10)
                    st.dataframe(changed_rows[['Tax Diff Status', 'Date Status', 'IGST Diff', 'CGST Diff', 'SGST Diff', 'Date Diff']], use_container_width=True)

# --- Merge Books Ledgers Page ---
if section == "Merge Books Ledgers":
    st.markdown("---")
    render_merge_ledgers_page()

# --- Performance Monitor Page ---
if section == "Performance Monitor":
    st.markdown("---")
    try:
        from utils.performance_dashboard import render_performance_dashboard_page
        render_performance_dashboard_page()
    except ImportError as e:
        st.error("Performance monitoring module not available. Please ensure all dependencies are installed.")
        st.code(str(e))
    except Exception as e:
        st.error(f"Error loading performance dashboard: {e}")
        st.info("The performance monitoring system requires additional setup. Please check the logs for more details.")

# --- Main Content Tabs (only show if not on Settings, Performance Monitor, or Merge Books Ledgers page) ---
if section not in ["Settings", "Performance Monitor", "Merge Books Ledgers"]:
    tabs = st.tabs(["Summary", "Transactions", "Visualizations", "Reports"])

    # --- Summary Tab ---
    with tabs[0]:
        st.markdown("#### Summary Reports")
        if st.session_state.get('reconciliation_complete') and st.session_state.get('reconciliation'):
            results = st.session_state.reconciliation.get_results()
            
            if st.button("Run Enhanced Reconciliation", key="enhanced_recon_btn"):
                with st.spinner("Running intelligent enhanced reconciliation..."):
                    # Save original summary for comparison
                    st.session_state.original_recon_summary = results.get('recon_summary').copy() if results.get('recon_summary') is not None else None
                    
                    # Run intelligent enhanced matching
                    analysis = st.session_state.reconciliation.run_intelligent_enhanced_matching()
                    
                    results = st.session_state.reconciliation.get_results()
                    final_report = results['final_report']
                    
                    # Apply settings to enhanced results
                    settings = get_current_settings()
                    if settings.get('auto_apply_settings', True):
                        final_report = apply_settings_to_reconciliation(final_report, settings)
                    
                    st.session_state.final_report = final_report
                    st.session_state.enhanced_run = True
                    
                    # Display analysis results
                    st.success("Intelligent enhanced reconciliation completed!")
                    
                    # Show analysis summary
                    st.markdown("##### Enhanced Matching Analysis")
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Books Only", analysis['books_only_count'])
                    with col2:
                        st.metric("GSTR-2A Only", analysis['gstr2a_only_count'])
                    with col3:
                        st.metric("Potential Groups", analysis['potential_gstin_inv_groups'])
                    
                    # Show recommendations
                    if analysis['recommendations']:
                        st.markdown("##### Matching Recommendations")
                        for i, rec in enumerate(analysis['recommendations'], 1):
                            st.markdown(f"**{i}. {rec['pattern']}** ({rec['priority']} priority)")
                            st.markdown(f"   {rec['description']}")
                            st.markdown(f"   Estimated matches: {rec['estimated_matches']}")
                    
                    st.rerun()
            
            # --- Raw Data Summary ---
            st.markdown("##### Raw Data Summary")
            raw_summary = results.get('raw_summary')
            if raw_summary is not None:
                st.dataframe(raw_summary.style.format("{:,.2f}", subset=pd.IndexSlice[:, ['Total IGST', 'Total CGST', 'Total SGST', 'Total Value']]), use_container_width=True)
            
            # --- Reconciliation Summary ---
            st.markdown("##### Reconciliation Summary")
            recon_summary = results.get('recon_summary')
            if recon_summary is not None:
                st.dataframe(recon_summary.style.format("{:,.2f}", subset=pd.IndexSlice[:, ['Total IGST', 'Total CGST', 'Total SGST', 'Total Value']]), use_container_width=True)
            
            # --- Enhanced Reconciliation Summary ---
            if st.session_state.get('enhanced_run') and st.session_state.get('original_recon_summary') is not None:
                st.markdown("##### Enhanced Reconciliation Summary (Comparison)")
                enhanced_summary = results.get('recon_summary')
                # Merge original and enhanced summaries for comparison
                if enhanced_summary is not None:
                    merged = st.session_state.original_recon_summary.copy()
                    merged.columns = [f"Before: {col}" if col != 'Match Type' else col for col in merged.columns]
                    enhanced_summary2 = enhanced_summary.copy()
                    enhanced_summary2.columns = [f"After: {col}" if col != 'Match Type' else col for col in enhanced_summary2.columns]
                    compare_df = pd.merge(merged, enhanced_summary2, on='Match Type', how='outer')
                    st.dataframe(compare_df, use_container_width=True)
            
            # --- Status Summary ---
            st.markdown("##### Status Summary")
            status_summary = results.get('status_summary')
            if status_summary is not None:
                st.dataframe(status_summary.style.format("{:,.2f}", subset=pd.IndexSlice[:, ['Matched Value', 'Mismatch Value', 'Total Value', 'Total IGST', 'Total CGST', 'Total SGST']]), use_container_width=True)
            
            # --- Detailed Match Summaries ---
            st.markdown("##### Detailed Match Summaries")
            col1, col2 = st.columns(2)
            summaries = st.session_state.reconciliation.get_summaries()
            with col1:
                st.write("###### Exact Matches Summary")
                st.dataframe(summaries.get('exact_matches_summary'), use_container_width=True)
            with col2:
                st.write("###### Partial Matches Summary")
                st.dataframe(summaries.get('partial_matches_summary'), use_container_width=True)

            col3, col4 = st.columns(2)
            with col3:
                st.write("###### Group Matches Summary")
                st.dataframe(summaries.get('group_matches_summary'), use_container_width=True)
            with col4:
                st.write("###### Cross-Year Matches Summary")
                st.dataframe(summaries.get('cross_year_matches_summary'), use_container_width=True)

            # Tax-wise summary (full width)
            st.markdown("##### Tax-wise Summary")
            st.dataframe(summaries.get('tax_summary'), use_container_width=True)
            
            # Add AI Insights section
            st.markdown("---")
            try:
                # Prepare reconciliation results for AI analysis
                reconciliation_results = {
                    "total_records": results.get('raw_summary', pd.DataFrame()).get('Count', pd.Series()).sum() if results.get('raw_summary') is not None else 0,
                    "matched_records": len(st.session_state.final_report[st.session_state.final_report['Status'].isin(['Exact Match', 'Partial Match', 'Group Match'])]) if st.session_state.final_report is not None else 0,
                    "processing_time": 0,  # Could be tracked if needed
                    "match_rate": 0,  # Could be calculated
                    "reconciliation_type": "gst_reconciliation"
                }
                
                # Calculate match rate
                if reconciliation_results["total_records"] > 0:
                    reconciliation_results["match_rate"] = (reconciliation_results["matched_records"] / reconciliation_results["total_records"]) * 100
                
                # Render AI insights section
                render_ai_insights_section_for_reports(reconciliation_results)
                
            except Exception as e:
                logger.error(f"Error adding AI insights to summary: {e}")
                # Fail silently to not disrupt main functionality
        else:
            st.info("Upload your Excel file and run reconciliation to see summary reports.")

    # --- Transactions Tab ---
    with tabs[1]:
        st.markdown("#### Transactions Table")
        # Create sub-tabs for Transactions, Unique GST Report, and GSTR-2A Compliance Report
        transaction_tabs = st.tabs(["Main Transactions", "Unique Unmapped GST Report", "GSTR-2A Compliance Report"])
        # Main Transactions sub-tab
        with transaction_tabs[0]:
            if st.session_state.get('reconciliation_complete') and st.session_state.get('final_report') is not None:
                df = st.session_state.final_report.copy()
                # --- Add compliance columns for GSTR-2A rows ---
                df = add_gstr2a_compliance_columns(df)
                required_cols = [
                    'IGST Diff', 'CGST Diff', 'SGST Diff', 'Date Diff', 'Status', 'Sub Status', 'Tax Diff Status',
                    'Date Status', 'Tax Sign Status', 'Value Sign', 'Narrative', 'Suggestions',
                    'Tax Head Status', 'GSTIN Match Status', 'Invoice Match Status', 'Trade Name Match Status', 'Legal Name Match Status',
                    'GSTIN Score', 'Trade Name Score', 'Compliance Report', 'Return Days Lapsed'
                ]
                for col in required_cols:
                    if col not in df.columns:
                        df[col] = np.nan
                df.insert(0, 'S.No.', range(1, len(df) + 1))
                df = df.astype(str)
                gb = GridOptionsBuilder.from_dataframe(df)
                gb.configure_default_column(editable=False, groupable=True, filter=True, sortable=True, resizable=True)
                for col in df.columns:
                    gb.configure_column(col, filter='agSetColumnFilter', resizable=True, sortable=True, minWidth=120)
                gb.configure_selection('multiple', use_checkbox=True)
                # --- AgGrid cellStyle for status and partial/group match highlighting ---
                status_cellstyle = JsCode('''
                    function(params) {
                        if (params.value == 'Exact Match') {
                            return { 'backgroundColor': '#43a047', 'color': 'white', 'fontWeight': 'bold' };
                        } else if (params.value == 'Partial Match') {
                            return { 'backgroundColor': '#fbc02d', 'color': '#222', 'fontWeight': 'bold' };
                        } else if (params.value == 'Group Match') {
                            return { 'backgroundColor': '#42a5f5', 'color': 'white', 'fontWeight': 'bold' };
                        } else if (params.value == 'Books Only') {
                            return { 'backgroundColor': '#1976d2', 'color': 'white', 'fontWeight': 'bold' };
                        } else if (params.value == 'GSTR-2A Only') {
                            return { 'backgroundColor': '#e53935', 'color': 'white', 'fontWeight': 'bold' };
                        } else if (params.value && params.value.toLowerCase().includes('cancellation')) {
                            return { 'backgroundColor': '#757575', 'color': 'white', 'fontWeight': 'bold' };
                        } else if (params.value) {
                            // Custom comments: light purple
                            return { 'backgroundColor': '#ab47bc', 'color': 'white', 'fontWeight': 'bold' };
                        }
                        return {};
                    }
                ''')
                # Highlight only the cells that are the reason for partial/group match
                partial_group_cellstyle = JsCode('''
                    function(params) {
                        var status = params.data['Status'];
                        var taxDiff = params.data['Tax Diff Status'];
                        var dateDiff = params.data['Date Status'];
                        var igstDiff = parseFloat(params.data['IGST Diff']);
                        var cgstDiff = parseFloat(params.data['CGST Diff']);
                        var sgstDiff = parseFloat(params.data['SGST Diff']);
                        var dateDiffVal = parseFloat(params.data['Date Diff']);
                        var invoiceNumberDiff = params.data['Invoice Number Diff'];
                        var taxTolerance = 0.01;
                        if (status == 'Partial Match' || status == 'Group Match' || status == 'Potential Match: Tax Deviation') {
                            // Highlight diff columns if not No Difference/Within Tolerance
                            if (
                                (params.colDef.field == 'IGST Diff' && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) ||
                                (params.colDef.field == 'CGST Diff' && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) ||
                                (params.colDef.field == 'SGST Diff' && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) ||
                                (params.colDef.field == 'Tax Diff Status' && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) ||
                                (params.colDef.field == 'Date Diff' && dateDiff == 'Outside Tolerance') ||
                                (params.colDef.field == 'Date Status' && dateDiff == 'Outside Tolerance')
                            ) {
                                return { 'backgroundColor': '#ffe0b2', 'fontWeight': 'bold' };
                            }
                            // Highlight only the original data columns where deviation occurred
                            if (params.colDef.field == 'Total IGST Amount' && Math.abs(igstDiff) > taxTolerance && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) {
                                return { 'backgroundColor': '#fffbe6', 'fontWeight': 'bold' };
                            }
                            if (params.colDef.field == 'Total CGST Amount' && Math.abs(cgstDiff) > taxTolerance && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) {
                                return { 'backgroundColor': '#fffbe6', 'fontWeight': 'bold' };
                            }
                            if (params.colDef.field == 'Total SGST Amount' && Math.abs(sgstDiff) > taxTolerance && (taxDiff == 'Has Difference' || taxDiff == 'High Deviation')) {
                                return { 'backgroundColor': '#fffbe6', 'fontWeight': 'bold' };
                            }
                            if (params.colDef.field == 'Invoice Date' && dateDiff == 'Outside Tolerance') {
                                return { 'backgroundColor': '#fffbe6', 'fontWeight': 'bold' };
                            }
                            if (params.colDef.field == 'Invoice Number' && invoiceNumberDiff == 'Yes') {
                                return { 'backgroundColor': '#fffbe6', 'fontWeight': 'bold' };
                            }
                        }
                        return {};
                    }
                ''')
                gb.configure_column('Status', cellStyle=status_cellstyle)
                # Apply partial/group match cellstyle to relevant columns
                for col in ['IGST Diff', 'CGST Diff', 'SGST Diff', 'Date Diff', 'Tax Diff Status', 'Date Status', 'Invoice Date', 'Invoice Number', 'Total IGST Amount', 'Total CGST Amount', 'Total SGST Amount']:
                    if col in df.columns:
                        gb.configure_column(col, cellStyle=partial_group_cellstyle)
                grid_options = gb.build()
                st.markdown("**Tip:** All Excel-like features are enabled: sort, filter, pin, autosize, choose columns, reset columns, etc. Use the column header menu!**")
                grid_response = AgGrid(
                    df,
                    gridOptions=grid_options,
                    theme='streamlit',
                    update_mode=GridUpdateMode.SELECTION_CHANGED | GridUpdateMode.VALUE_CHANGED,
                    allow_unsafe_jscode=True,
                    fit_columns_on_grid_load=True,
                    enable_enterprise_modules=True,
                    height=600,
                    width='100%'
                )
                selected = grid_response['selected_rows']
                # Bulk status update
                st.markdown("**Bulk Update Status**: Select rows and choose a new status/comment to apply.")
                new_status = st.selectbox("Set Status to", ["Exact Match", "Partial Match", "Books Only", "GSTR-2A Only", "Custom Comment"], key="bulk_status_select")
                custom_comment = ""
                if new_status == "Custom Comment":
                    custom_comment = st.text_input("Enter custom comment", key="bulk_custom_comment")
                if st.button("Apply to Selected Rows", key="bulk_apply_status") and selected:
                    for row in selected:
                        idx = int(row['S.No.']) - 1
                        if new_status == "Custom Comment":
                            df.at[idx, "Status"] = custom_comment
                        else:
                            df.at[idx, "Status"] = new_status
                    st.session_state.final_report = df
                    st.success(f"Updated status for {len(selected)} rows.")
                    st.rerun()
                st.info("Scroll left/right and up/down to view all columns. All original and reconciliation columns are included. Use column headers to filter and sort. You can also resize and reorder columns. Table header is sticky and always visible.")
            else:
                st.info("Run reconciliation to view transactions table.")
        # Unique GST Report sub-tab
        with transaction_tabs[1]:
            if st.session_state.get('reconciliation_complete') and st.session_state.get('final_report') is not None:
                render_unique_gst_report(st.session_state.final_report)
            elif st.session_state.get('df') is not None:
                st.info("📋 Please run reconciliation first to generate the Unique GST Report.")
                st.markdown("""
                ### About the Unique GST Report
                
                This report identifies GST numbers from GSTR-2A that have **'GSTR-2A Only' status** (no corresponding entries in your Books data).
                
                **What it shows:**
                - GST numbers that appear only in GSTR-2A (not in Books)
                - Associated Trade Names and Legal Names
                - Count of occurrences and total tax amounts
                - Summary statistics and insights
                
                **Use cases:**
                - Identify missing supplier records in your books
                - Find potential data entry errors
                - Audit supplier master data completeness
                - Assess financial impact of unmapped suppliers
                
                **Features:**
                - 🔍 Search and filter functionality
                - 📊 Interactive table with proper sorting (ascending/descending)
                - 📥 Export to CSV/Excel
                - 📈 Detailed insights and analytics
                - 💰 Indian currency formatting (lakhs/crores)
                """)
            else:
                st.info("📋 Please upload your Excel file first to generate the Unique GST Report.")
                st.markdown("""
                ### About the Unique GST Report
                
                This report identifies GST numbers from GSTR-2A that have **'GSTR-2A Only' status** (no corresponding entries in your Books data).
                
                **What it shows:**
                - GST numbers that appear only in GSTR-2A (not in Books)
                - Associated Trade Names and Legal Names
                - Count of occurrences and total tax amounts
                - Summary statistics and insights
                
                **Use cases:**
                - Identify missing supplier records in your books
                - Find potential data entry errors
                - Audit supplier master data completeness
                - Assess financial impact of unmapped suppliers
                
                **Features:**
                - 🔍 Search and filter functionality
                - 📊 Interactive table with proper sorting (ascending/descending)
                - 📥 Export to CSV/Excel
                - 📈 Detailed insights and analytics
                - 💰 Indian currency formatting (lakhs/crores)
                """)
        # GSTR-2A Compliance Report sub-tab
        with transaction_tabs[2]:
            st.markdown("#### GSTR-2A Compliance Report")
            if st.session_state.get('reconciliation_complete') and st.session_state.get('final_report') is not None:
                if 'compliance_report_generated' not in st.session_state:
                    st.session_state.compliance_report_generated = False
                if st.button("Generate Compliance Report", key="generate_compliance_report") or st.session_state.compliance_report_generated:
                    st.session_state.compliance_report_generated = True
                    with st.spinner("Generating compliance summary and visualizations..."):
                        try:
                            summary_df = get_gstr2a_compliance_summary(st.session_state.final_report)
                            if summary_df.empty:
                                st.info("No GSTR-2A data available for compliance report.")
                            else:
                                st.markdown("##### Compliance Summary Table (GSTIN-wise, Month-wise)")
                                gb = GridOptionsBuilder.from_dataframe(summary_df)
                                gb.configure_default_column(editable=False, groupable=True, filter=True, sortable=True, resizable=True)
                                for col in summary_df.columns:
                                    gb.configure_column(col, filter='agSetColumnFilter', resizable=True, sortable=True, minWidth=120)
                                grid_options = gb.build()
                                grid_response = AgGrid(
                                    summary_df,
                                    gridOptions=grid_options,
                                    theme='streamlit',
                                    update_mode=GridUpdateMode.SELECTION_CHANGED,
                                    allow_unsafe_jscode=True,
                                    fit_columns_on_grid_load=True,
                                    enable_enterprise_modules=True,
                                    height=400,
                                    width='100%'
                                )
                                st.info("Tip: All Excel-like features are enabled: sort, filter, pin, autosize, choose columns, reset columns, etc. Use the column header menu!")
                                # --- Export options ---
                                output = io.BytesIO()
                                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                    summary_df.to_excel(writer, sheet_name='Compliance Summary', index=False)
                                excel_data = output.getvalue()
                                st.download_button(
                                    label="⬇️ Download Compliance Summary (Excel)",
                                    data=excel_data,
                                    file_name=f"gstr2a_compliance_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                                csv_data = summary_df.to_csv(index=False).encode('utf-8')
                                st.download_button(
                                    label="⬇️ Download Compliance Summary (CSV)",
                                    data=csv_data,
                                    file_name="gstr2a_compliance_summary.csv",
                                    mime="text/csv",
                                )
                                # --- Visualizations ---
                                st.markdown("##### Visualizations")
                                # Filters
                                gstin_list = summary_df['Supplier GSTIN'].unique().tolist()
                                months = [col for col in summary_df.columns if col not in ['Supplier GSTIN', 'Trade Name', 'Legal Name', 'Total Risky Months', 'Total Compliant Months']]
                                selected_gstin = st.multiselect("Filter by GSTIN", gstin_list, default=gstin_list[:5] if len(gstin_list) > 5 else gstin_list)
                                selected_months = st.multiselect("Filter by Month", months, default=months)
                                filtered_df = summary_df[summary_df['Supplier GSTIN'].isin(selected_gstin)]
                                # Heatmap: Compliance status (Compliant/Risky) across months for top suppliers
                                st.markdown("**Heatmap: Compliance Status Across Months (Top Suppliers)**")
                                import plotly.graph_objects as go
                                heatmap_data = filtered_df.set_index('Supplier GSTIN')[selected_months]
                                z = heatmap_data.applymap(lambda x: 1 if isinstance(x, str) and x.startswith('Risky') else (0 if x == 'Compliant' else np.nan)).values
                                fig = go.Figure(data=go.Heatmap(
                                    z=z,
                                    x=selected_months,
                                    y=filtered_df['Supplier GSTIN'],
                                    colorscale=[[0, 'green'], [1, 'red']],
                                    colorbar=dict(title='Risky=1, Compliant=0'),
                                    hoverinfo='text',
                                    text=heatmap_data.values
                                ))
                                st.plotly_chart(fig, use_container_width=True)
                                st.caption("Red: Risky, Green: Compliant. Blank: No data.")
                                # Bar Chart: Number of Risky vs. Compliant months per supplier
                                st.markdown("**Bar Chart: Risky vs. Compliant Months per Supplier**")
                                bar_df = filtered_df[['Supplier GSTIN', 'Total Risky Months', 'Total Compliant Months']].set_index('Supplier GSTIN')
                                fig2 = go.Figure()
                                fig2.add_bar(x=bar_df.index, y=bar_df['Total Risky Months'], name='Risky Months', marker_color='red')
                                fig2.add_bar(x=bar_df.index, y=bar_df['Total Compliant Months'], name='Compliant Months', marker_color='green')
                                fig2.update_layout(barmode='group', xaxis_title='Supplier GSTIN', yaxis_title='Months')
                                st.plotly_chart(fig2, use_container_width=True)
                                # Line Chart: Avg Return Days Lapsed per month for Risky suppliers
                                st.markdown("**Line Chart: Avg Return Days Lapsed per Month (Risky Suppliers)**")
                                # For this, need to process the main table for risky filings
                                main_df = st.session_state.final_report.copy()
                                main_df = add_gstr2a_compliance_columns(main_df)
                                main_df['Month'] = pd.to_datetime(main_df['Invoice Date'], errors='coerce').dt.strftime('%b %Y')
                                risky_main = main_df[main_df['Compliance Report'].str.startswith('Risky', na=False)]
                                avg_days = risky_main.groupby('Month')['Return Days Lapsed'].apply(lambda x: pd.to_numeric(x, errors='coerce').mean()).reindex(months)
                                fig3 = go.Figure()
                                fig3.add_trace(go.Scatter(x=months, y=avg_days.values, mode='lines+markers', name='Avg Return Days Lapsed', line=dict(color='orange')))
                                fig3.update_layout(xaxis_title='Month', yaxis_title='Avg Return Days Lapsed')
                                st.plotly_chart(fig3, use_container_width=True)
                                # Pie Chart: Proportion of Risky vs. Compliant transactions by tax amount
                                st.markdown("**Pie Chart: Proportion of Risky vs. Compliant Transactions by Tax Amount**")
                                risky_amt = risky_main['IGST'].astype(float).fillna(0).sum() + risky_main['CGST'].astype(float).fillna(0).sum() + risky_main['SGST'].astype(float).fillna(0).sum()
                                compliant_main = main_df[main_df['Compliance Report'] == 'Compliant']
                                compliant_amt = compliant_main['IGST'].astype(float).fillna(0).sum() + compliant_main['CGST'].astype(float).fillna(0).sum() + compliant_main['SGST'].astype(float).fillna(0).sum()
                                fig4 = go.Figure(data=[
                                    go.Pie(labels=['Risky', 'Compliant'], values=[risky_amt, compliant_amt], hole=0.4)
                                ])
                                st.plotly_chart(fig4, use_container_width=True)
                                st.caption("Pie shows total tax amount for Risky vs. Compliant filings.")
                        except Exception as e:
                            import logging
                            logging.exception("Error generating compliance report")
                            st.error(f"Error generating compliance report: {e}")
            else:
                st.info("Run reconciliation to view GSTR-2A Compliance Report.")

    # --- Raw Data Preview (add S.No. column) ---
    if st.session_state.uploaded_file is not None and st.session_state.edited_df is not None:
        raw_df = st.session_state.edited_df.copy()
        if 'S.No.' not in raw_df.columns:
            raw_df.insert(0, 'S.No.', range(1, len(raw_df) + 1))
        st.subheader("Raw Data Preview")
        st.dataframe(raw_df, use_container_width=True)

    # --- Visualizations Tab ---
    with tabs[2]:
        st.markdown("#### Visualizations")
        st.info("Charts and analytics coming soon!")

    # --- Reports Tab (for downloads) ---
    with tabs[3]:
        st.markdown("#### Download Reports")
        if st.session_state.get('reconciliation_complete') and st.session_state.get('final_report') is not None:
            final_report_df = st.session_state.final_report.copy()
            # --- Excel Download ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add compliance columns to final report before export
                final_report_df = add_gstr2a_compliance_columns(final_report_df)
                final_report_df.to_excel(writer, sheet_name='Final Report', index=False)
                results = st.session_state.reconciliation.get_results()
                if results.get('raw_summary') is not None:
                    results['raw_summary'].to_excel(writer, sheet_name='Raw Data Summary', index=False)
                if results.get('recon_summary') is not None:
                    results['recon_summary'].to_excel(writer, sheet_name='Reconciliation Summary', index=False)
                if results.get('status_summary') is not None:
                    results['status_summary'].to_excel(writer, sheet_name='Status Summary', index=False)
                if results.get('exact_matches_summary') is not None:
                    results['exact_matches_summary'].to_excel(writer, sheet_name='Exact Matches Summary', index=False)
                if results.get('partial_matches_summary') is not None:
                    results['partial_matches_summary'].to_excel(writer, sheet_name='Partial Matches Summary', index=False)
                if results.get('group_matches_summary') is not None:
                    results['group_matches_summary'].to_excel(writer, sheet_name='Group Matches Summary', index=False)
                if results.get('tax_summary') is not None:
                    results['tax_summary'].to_excel(writer, sheet_name='Tax Summary', index=False)
                # Add Unique Unmapped GST Numbers summary
                unmapped_gst_summary = get_unmapped_gst_summary_for_excel(st.session_state.final_report)
                if not unmapped_gst_summary.empty:
                    unmapped_gst_summary.to_excel(writer, sheet_name='Unique Unmapped GST Numbers', index=False)
                # Add GSTR-2A Compliance Report summary as a new sheet
                compliance_summary = get_gstr2a_compliance_summary(final_report_df)
                if compliance_summary is not None and not compliance_summary.empty:
                    compliance_summary.to_excel(writer, sheet_name='GSTR-2A Compliance Report', index=False)
                # --- Apply color formatting to Final Report sheet ---
                workbook = writer.book
                ws = workbook['Final Report']
                # Status color mapping (for Status column only)
                status_colors = {
                    'Exact Match': '43a047',
                    'Partial Match': 'fbc02d',
                    'Group Match': '42a5f5',
                    'Books Only': '1976d2',
                    'GSTR-2A Only': 'e53935',
                    'Sign Cancellation': '757575',
                }
                # Diff highlight colors
                diff_fill = PatternFill(start_color='ffe0b2', end_color='ffe0b2', fill_type='solid')  # orange
                orig_diff_fill = PatternFill(start_color='fffbe6', end_color='fffbe6', fill_type='solid')  # light yellow
                # Find column indices
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                status_idx = header.index('Status') + 1 if 'Status' in header else None
                tax_diff_idx = header.index('Tax Diff Status') + 1 if 'Tax Diff Status' in header else None
                date_status_idx = header.index('Date Status') + 1 if 'Date Status' in header else None
                igst_diff_idx = header.index('IGST Diff') + 1 if 'IGST Diff' in header else None
                cgst_diff_idx = header.index('CGST Diff') + 1 if 'CGST Diff' in header else None
                sgst_diff_idx = header.index('SGST Diff') + 1 if 'SGST Diff' in header else None
                date_diff_idx = header.index('Date Diff') + 1 if 'Date Diff' in header else None
                total_igst_idx = header.index('Total IGST Amount') + 1 if 'Total IGST Amount' in header else None
                total_cgst_idx = header.index('Total CGST Amount') + 1 if 'Total CGST Amount' in header else None
                total_sgst_idx = header.index('Total SGST Amount') + 1 if 'Total SGST Amount' in header else None
                invoice_date_idx = header.index('Invoice Date') + 1 if 'Invoice Date' in header else None
                invoice_number_idx = header.index('Invoice Number') + 1 if 'Invoice Number' in header else None
                # Apply cell coloring
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    status = row[status_idx-1].value if status_idx else None
                    # Only color the Status cell
                    if status and status_idx:
                        for key, color in status_colors.items():
                            if key in status:
                                row[status_idx-1].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                                break
                    # Highlight diff columns for partial/group/potential
                    if status in ['Partial Match', 'Group Match', 'Potential Match: Tax Deviation']:
                        # Tax diff highlight (orange)
                        if tax_diff_idx and row[tax_diff_idx-1].value in ['Has Difference', 'High Deviation']:
                            for idx in [igst_diff_idx, cgst_diff_idx, sgst_diff_idx, tax_diff_idx]:
                                if idx:
                                    row[idx-1].fill = diff_fill
                        # Date diff highlight (light yellow)
                        if date_status_idx and row[date_status_idx-1].value == 'Outside Tolerance':
                            for idx in [date_diff_idx, date_status_idx]:
                                if idx:
                                    row[idx-1].fill = orig_diff_fill
                        # Also highlight original data columns for tax/date diff (light yellow)
                        if tax_diff_idx and row[tax_diff_idx-1].value in ['Has Difference', 'High Deviation']:
                            for idx in [total_igst_idx, total_cgst_idx, total_sgst_idx]:
                                if idx:
                                    row[idx-1].fill = orig_diff_fill
                        if date_status_idx and row[date_status_idx-1].value == 'Outside Tolerance':
                            if invoice_date_idx:
                                row[invoice_date_idx-1].fill = orig_diff_fill
                # End color formatting
            excel_data = output.getvalue()
            st.download_button(
                label="⬇️ Download All Reports (Excel)",
                data=excel_data,
                file_name="gst_reconciliation_reports.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel_reports"
            )
            # --- CSV Download ---
            csv_data = final_report_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="⬇️ Download Final Report (CSV)",
                data=csv_data,
                file_name="gst_reconciliation_report.csv",
                mime="text/csv",
                key="download_csv_report"
            )
        else:
            st.info("Run reconciliation to generate and download reports.")

    # --- Post-Reconciliation GSTIN Comments Module ---
    st.markdown("---")
    st.markdown("## Post-Reconciliation GSTIN Comments")
    st.markdown("Configure comments to be applied to specific GSTINs in the final report. These comments will update the Status column for matching records.")
    
    # Show confirmation message if present (outside form so it persists after rerun)
    if 'gstin_comment_message' in st.session_state and st.session_state.gstin_comment_message:
        msg, msg_type = st.session_state.gstin_comment_message
        if msg_type == 'success':
            st.success(msg)
        elif msg_type == 'info':
            st.info(msg)
        st.session_state.gstin_comment_message = None

    # Initialize sets if not present
    if 'gstin_comment_sets' not in st.session_state:
        st.session_state.gstin_comment_sets = [
            {'gstins': '', 'comment': '', 'status': 'Both'}
        ]
    status_options = [
        'Both', 'Exact Match', 'Partial Match', 'Books Only', 'GSTR-2A Only', 'Custom Comment'
    ]

    # --- Add/Remove Set Buttons (OUTSIDE the form) ---
    col_add, col_remove = st.columns([1, 3])
    with col_add:
        if st.button("Add Another GSTIN Set", key="add_gstin_set"):
            st.session_state.gstin_comment_sets.append({'gstins': '', 'comment': '', 'status': 'Both'})
            st.rerun()
    with col_remove:
        if len(st.session_state.gstin_comment_sets) > 1:
            for i in range(len(st.session_state.gstin_comment_sets)):
                remove_label = f"Remove Set {i+1}"
                if st.button(remove_label, key=f"remove_gstin_set_{i}"):
                    st.session_state.gstin_comment_sets.pop(i)
                    st.rerun()

    # --- FORM START ---
    with st.form("gstin_comments_form"):
        new_sets = []
        for i, set_ in enumerate(st.session_state.gstin_comment_sets):
            with st.expander(f"GSTIN Comment Set {i+1}", expanded=True):
                gstins_val = st.text_area(f"Enter GSTINs (one per line)", value=set_['gstins'], key=f'gstins_{i}')
                comment_val = st.text_input(f"Comment", value=set_['comment'], key=f'comment_{i}')
                status_val = st.selectbox(f"Apply to Status", options=status_options, index=status_options.index(set_['status']) if set_['status'] in status_options else 0, key=f'status_{i}')
                new_sets.append({'gstins': gstins_val, 'comment': comment_val, 'status': status_val})
        # --- Summary Table ---
        st.markdown("### Summary of Post-Reconciliation GSTIN Comment Sets")
        for i, set_ in enumerate(new_sets):
            gstin_list = [g.strip().upper() for g in set_['gstins'].splitlines() if g.strip()]
            st.write(f"Set {i+1}: {len(gstin_list)} GSTINs, Comment: '{set_['comment']}', Status: {set_['status']}")
        # --- Apply Comments Button ---
        apply_comments = st.form_submit_button("Apply Comments")
        if apply_comments:
            updated_df = st.session_state.final_report.copy()
            total_updated = 0
            for set_ in new_sets:
                gstin_list = [g.strip().upper() for g in set_['gstins'].splitlines() if g.strip()]
                comment = set_['comment']
                status = set_['status']
                if not gstin_list or not comment:
                    continue
                mask_gstin = updated_df['Supplier GSTIN'].astype(str).str.upper().isin(gstin_list)
                if status == 'Both':
                    mask_status = mask_gstin
                else:
                    mask_status = mask_gstin & (updated_df['Status'] == status)
                n_updated = mask_status.sum()
                total_updated += n_updated
                updated_df.loc[mask_status, 'Status'] = comment
            st.session_state.final_report = updated_df
            st.session_state.gstin_comment_sets = new_sets
            if total_updated > 0:
                st.session_state.gstin_comment_message = (f"✅ Comments applied successfully! {total_updated} rows updated.", 'success')
            else:
                st.session_state.gstin_comment_message = ("ℹ️ No rows matched the criteria. No comments were applied.", 'info')
            st.rerun()
        else:
            st.session_state.gstin_comment_sets = new_sets

if __name__ == "__main__":
    initialize_session_state() 